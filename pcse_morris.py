# %%
import numpy as np
from SALib.sample.morris import sample
from SALib.analyze.morris import analyze

from dmitte import para_constant, run_wofost
from dmitte.transfer_equotions import transfer_rates, solve_con, transfer_rates_UFORTI, transfer_rates_baomi
from dmitte.para_constant import LAMBDA_T_H
from openpyxl import load_workbook
from joblib import Parallel, delayed
from tqdm import tqdm
import uuid

from plot_config import *

def meteodata(start_date,end_date, params): 
    df = pd.read_excel('./data/meteo/meteo_usefor_cttm.xlsx')
    df['TIMESTAMP'] = pd.to_datetime(df['TIMESTAMP'])
    df.set_index('TIMESTAMP', inplace=True)

    full_range = pd.date_range(start=start_date, end=end_date, freq='D')
    df = df.reindex(full_range)

    df.fillna(method='bfill', inplace=True)  # 先使用后一天的数据填充
    df.fillna(method='ffill', inplace=True)  # 然后使用前一天的数据填充

    # 得到指定日期范围的数据
    filtered_df = df.loc[start_date:end_date]   
    # 使用线性插值填充缺失值
    hourly_df = filtered_df.resample('H').interpolate(method='linear')

    names = ['WS10m_avg', 'Ta_Avg', 'DR_Avg', 'Rain_Tot', 'RH_Avg', 'Pvapor_Avg', 'VWC_5cm_Avg', 'VWC_10cm_Avg', 'VWC_20cm_Avg']
    for i, name in enumerate(names):
        hourly_df[name] = float(params[i])

    return hourly_df

class Calc_air():
    '''
    calc_washout()      计算湿沉降速率
    calc_USTERN ()      计算摩擦速度
    calc_row()          计算饱和水蒸汽含量
    calc_airesistance() 计算RAM和RB大气阻滞因子、边界层阻滞因子
    calc_res()          调用本class中所有函数
    '''
    def __init__(self, pcse_results, start_date, end_date, plant_type, sta, HEG, x, params):
        self.meteo = meteodata(start_date,end_date, params)
        self.x  = x                            # 下风向距离,m
        self.sta = sta             # 稳定度分类
        self.heg = HEG                         # 扩散高度,m,10、30、50
        self.ICOMP = para_constant.Plant_dict[plant_type.upper()]['ICOMP']      # 1表示农业,2表示牧草模型, 3土豆,4小麦  
        self.U10 = np.array (self.meteo['WS10m_avg'], dtype=np.float64)         # 10m风速, m/s
        self.TEMP = np.array (self.meteo['Ta_Avg'], dtype=np.float64)           # 实际温度, ℃
        self.PAR = np.array (self.meteo['DR_Avg'], dtype=np.float64)            # 太阳辐射, W/m2
        rainfall = np.array (self.meteo['Rain_Tot'], dtype=np.float64)          # 降雨量, mm/h;
        self.rainfall = rainfall / 24
        RH = np.array (self.meteo['RH_Avg'], dtype=np.float64)                  # 相对湿度, % 
        self.r_h = RH / 100
        self.ea = np.array (self.meteo['Pvapor_Avg'], dtype=np.float64)         # 实际蒸汽压,kpa

        LAI = pcse_results['LAI']     # LAI :叶片指数, m2/m2
        self.LAI = LAI.values

        TWLV = pcse_results['TWLV']   # TWLV:全部叶片的干重, kg/ha
        tagp = TWLV.values
        self.BWG = tagp * 1E-5        # BWG: 单位面积叶片含水量,g/cm-2

        WWLOW = pcse_results['WWLOW']  # WWLOW:土壤水含量,cm
        waterinsoil = WWLOW.values  
        self.BODWA = waterinsoil * 10  # 土壤实际含水量,kg m-2 

        self.calc_res()
        
    # 湿沉积修正
    def calc_washout(self):
        # rainfall 雨量, 小雨 <1mm/h,中雨 1-3mm/h,大雨>3mm/h ；
        # 湿沉积因子 Washout coefficient, s-1
        # washout_coeff = np.where(x > 300,
        #                          np.where(self.rainfall > 3, 3.4e-4, np.where(self.rainfall >= 1, 1.22e-4, np.where(self.rainfall > 0, 2.9e-5, 0))),
        #                          np.where(self.rainfall > 3, 4.0e-4, np.where(self.rainfall >= 1, 2.5e-5, np.where(self.rainfall > 0, 1.0e-4, 0))))
        washout_coeff = 9E-5 * (self.rainfall ** 0.6)
        self.Fwet = np.exp(-washout_coeff * self.x / self.U10)

    # 计算摩擦速度,USTERN
    def calc_USTERN(self, IROUGH=2, IRAU=0.1):
        '''
        IROUGH:  粗糙度类型, 取值为1、2、3;
        IRAU:    粗糙度Z0,  IROUGH=2为平均粗糙度(低矮植物,乡村地区,粗糙度Z0的大小为10cm-1m);
                        IROUGH=3为大粗糙度(森林,城市地区,粗糙度Z0的大小为>1m)
                        ! zo [m]   description
                            0.01     lawn grass, bodies of water
                            0.04     plowed land
                            0.10     open grassland
                            0.40     rural areas, mixed farming, woods, small villages
                            1.00     cities and forests
                            4.00     cities with tall buildings
        '''
        A0 = para_constant.A0
        B0 = para_constant.B0
        MEX = para_constant.ISK            # 风廓线幂指数值
        RLMON = A0[self.sta - 1] * IRAU ** B0[self.sta - 1]
        self.U100 = self.U10 * (10 / 100) ** MEX[self.sta - 1]

        if self.sta == 4:  # sta=4,表示中性条件
            A10 = 60.0 - (6.0 * IRAU)
            A11 = np.log (A10 / IRAU)
            self.USTERN = (self.U100 * 0.41) / A11

        elif self.sta <= 3:  # sta<=3,表示不稳定条件
            L = 1.0 / RLMON
            # L表示 MONIN-OBUCHOW 长度,  < 0 UNSTABLE ; > 0 STABLE
            AN = -0.25
            A10 = 100 - 6.0 * IRAU
            PHIM = (1 - (15.0 * (A10) / L)) ** AN
            A11 = 0.5 * (1.0 + (1.0 / PHIM))
            A12 = np.log(A11)
            A21 = 0.5 * (1.0 + (1.0 / PHIM ** 2))
            A22 = np.log(A21)
            A31 = 1.0 / PHIM
            A32 = 2 * np.arctan(A31)
            A41 = np.log((A10) / IRAU)
            self.USTERN = self.U100 * 0.41 / (A41 - 2.0 * A12 - A22 + 2.0 * A32 - (np.pi / 2))

        elif self.sta == 5:  # sta=5,表示稳定条件
            if IROUGH <= 2:
                URN = 20.0
            elif IROUGH == 3:
                URN = 30.0
            # U100 = self.U10 * (URN / self.ZREF) ** MEX[self.sta - 1]
            U100 = self.U10 * (URN / 100) ** MEX[self.sta - 1]
            L = 1.0 / RLMON
            A31 = np.log((URN - 6.0 * IRAU) / IRAU)
            self.USTERN = U100 * 0.41 / (A31 + (5.0 * (URN - 6.0 * IRAU) / L))

        elif self.sta == 6:  # sta=6,表示稳定条件
            if IROUGH <= 2:
                URN = 10.0
            elif IROUGH == 3:
                URN = 20.0
            U100 = self.U10 * (URN / 100) ** MEX[self.sta - 1]
            L = 1.0 / RLMON
            A31 = np.log((URN - 6.0 * IRAU) / IRAU)
            self.USTERN = U100 * 0.41 / (A31 + (5.0 * (URN - 6.0 * IRAU) / L))
    
    # 计算 饱和水蒸汽含量,ROW g/m3
    def calc_row(self):
        AMW = 18.016     # 水的分子量,g/mol
        RAW = 8.3141E03  # 气体常数,J/(mol·K)
        AM1 = (7.5 * self.TEMP) / (237. + self.TEMP)
        AM2 = 6.107 * 10. ** (AM1)
        self.ROW = (AMW * AM2 * 1.E5) / (RAW * (self.TEMP + 273.))
    
    # 计算大气气动阻力因子,RAM, s/m;
    # 计算大气边界阻力因子,RB, s/m;
    def calc_airesistance(self):
        RAM = (self.U100 / self.USTERN ** 2)
        self.RAM = RAM
        k = 0.4  # karman常数;
        RB = 2.0 / (k * self.USTERN)
        self.RB = RB
        # self.LAI = np.where (self.LAI > 1, self.LAI / self.RB, self.LAI)
    
    # 计算大气水库室中的氢含量atm_h 和 大气库室中总空气的含量atm_a;
    def calc_aircontent(self):
        '''
        ML 混合层高度,m;
        temp 空气温度(℃);
        r_h 相对湿度,%;
        a_h 绝对湿度,g/m3;
        atm_pressure 大气压,hpa(百帕)  e_s 饱和水蒸气压
        '''
        atm_p = np.array (self.meteo['P_Avg'], dtype=np.float64)
        e_s = 6.112 * np.exp ((17.62 * self.TEMP) / (self.TEMP + 243.12))  # 麦尔宁方程(Magnus formula)
        fp = 1.0016 + 3.15e-6 * atm_p - 0.074 / atm_p
        self.a_h = 1000 * (self.r_h *100) * e_s * fp / (461.5 * (self.TEMP + 273.15))  # 461.5是水蒸气的特定气体常数
        # print("a_h",a_h)
        self.ML = para_constant.mixLayerH[self.sta - 1]
        self.atm_w = (self.ML * self.a_h / 1000) * 1E6   # 空气中水的量, kg/ km2
        self.atm_h = (self.ML * self.a_h / 1000) * 0.11  # 0.11是将水的含量转换为氢含量,kg/m2
        self.atm_a = self.ML * 1e6                       # 1e6 是 1km2;

    def calc_res(self):
        self.calc_washout()
        self.calc_USTERN ()
        self.calc_row()
        self.calc_airesistance()
        self.calc_aircontent()

class Calc_soil(Calc_air):
    '''
    calc_RSOIL ()    计算土壤阻滞因子
    calc_VDSO()      计算土壤沉积速率
    calc_ESOIL()     计算土壤再释放率
    calc_soil_ab()   计算土壤水分迁移速率
    calc_res()       调用本class中所有函数
    '''
    # pcse_results, start_date, end_date, plant_type, sta, HEG, x
    def __init__(self, pcse_results, start_date, end_date, plant_type, sta, HEG, x, params):
        super ().__init__ (pcse_results, start_date, end_date, plant_type, sta, HEG, x, params)
        self.BODW1 = np.array (self.meteo['VWC_5cm_Avg'])
        self.calc_soilfx()

    # 计算 土壤阻力因子,RSOIL, s/m
    def calc_RSOIL(self, thetMAX_i=0.5):
        '''
        thetMAX_i: 第一层土壤最大含水量,%
        BODW1:  第一层土壤实际含水量,%
        ZSOIL:  干土层深度,m
        '''
        tort = 1.5             # 土壤空隙的弯曲度因子
        RSSMIN = 200           # 最小土壤阻滞因子,s/m, 范围100-300之间
        D0 = 0.23E-4           # 空气中水汽扩散系数, m2/s
        z0 = 0.004             # 干土层的初始厚度,m
        phi_h = 0.5            # 第一cm表层土的实际含水量,范围在0.5-5之间
        PHI = (thetMAX_i - self.BODW1) / self.BODW1
        PHI = PHI * np.ones_like (self.TEMP)
        TEMPA = np.maximum (self.TEMP + 273.15, 0)  # 将负数值替换为零
        Deff = D0 * (TEMPA / 273.15) ** 1.75
        ZSOIL = z0 / phi_h                          # phi 实际土壤含水量
        ZSOIL = ZSOIL * np.ones_like (self.TEMP)
        self.RSOIL = ZSOIL * PHI / (Deff * tort) + RSSMIN
        self.RSOIL = np.where (self.rainfall > 10, self.RSOIL / 3, self.RSOIL)
        self.RSOIL = np.where ((self.rainfall <= 10) & (self.r_h > 0.98), self.RSOIL / 3, self.RSOIL)

    # 计算土壤沉积速率 VDSO, m/s
    def calc_VDSO(self):
        self.VDSO = 1 / (self.RAM + self.RB + self.RSOIL)

    # 计算 土壤再蒸发率 REEM,kg m-2 h-1
    def calc_ESOIL(self):
        """
        计算 土壤再蒸发率 REEM, kg m-2 h-1

        RELFEU (float): 相对湿度 (0-1)
        TEMP (float): 温度 (摄氏度)
        PAR (float): 光合有效辐射 (W/m^2)
        LEAFA (float): 叶面积指数 (m^2/m^2)
        """
        RATMOS = self.RAM + self.RB             # 大气阻滞因子之和,单位 m/s
        RSOILA = self.RSOIL                     # 土壤阻滞因子,m/s
        es = 0.6108 * np.exp(17.27 * self.TEMP / (self.TEMP + 237.3))  # 饱和蒸汽压,kpa,Tetens公式
        EFEUDE = (es - self.ea * self.r_h)         # 空气实际与饱和的蒸汽压只差,N m-2
        DELE = 4098 * es / (self.TEMP + 237.3)**2    # 饱和水汽压的温度依赖,单位 Pa
        
        GAM = 0.67                            # 大气干湿度压力系数
        RHO = 1.2923                          # 空气密度,单位 kg/m^3
        CP = 1.013                            # 比热容,单位 J/(kg K)
        LE = 2.45E+6                          # 水的蒸发潜热,单位 J/kg
        ALBEDO = 0.23                         # 地表反射率

        # Rno = 2. * (1. - ALBEDO) * self.PAR
        self.QSTRS = self.PAR * np.exp(-0.398 * self.LAI)  # 到达土壤的辐射通量
        ESOIL = ((DELE * self.QSTRS) + (RHO * CP * EFEUDE / RATMOS)) / (DELE + GAM * (1. + RSOILA / RATMOS)) / LE
        self.ESOIL = ESOIL * 3600             # 此处,土壤实际蒸发率单位kg m-2 h-1；
        

    # 计算 土壤的水分迁移速率 Va_b , m h-1
    def calc_soil_ab(self, thetMAX_layers=[0.5, 0.6, 0.65]):
        """
        Calculates the moisture migration rate, Va_b, between two soil layers.

        theta: a is the volumetric water content of layer a, b is that of layer b, m3 m-3;
        theta_s_layers: Saturated soil water content for each layer, m3 m-3;
        theta_33_layers: Field capacity for each layer, m3 m-3;
        theta_1500_layers: Permanent wilting point for each layer, m3 m-3;
        thetMAX_layers: Maximum soil water content for each layer, layer 1 is 0.5, layer 2 is 0.35
        """
        theta = np.array (self.meteo[['VWC_5cm_Avg', 'VWC_10cm_Avg', 'VWC_20cm_Avg']])
        theta_s_layers = 0.4 * np.ones_like(theta[:,0])
        theta_33_layers = 0.27 * np.ones_like(theta[:,0])
        theta_1500_layers = 0.09 * np.ones_like(theta[:,0])

        def calc_SS(thetMAX_i, theta_i, theta_1500_i):
            PSI = (thetMAX_i - theta_i) / (thetMAX_i - theta_1500_i)
            SS = 1.5e5 * PSI ** (1.65 + PSI * 7.3 + (-3.1) * PSI ** 7.5)
            return SS

        def calc_permeability_k(theta_i, theta_s_i, theta_33_i, theta_1500_i):

            """
            计算土壤的渗透系数 K, m d-1。

            参数:
            theta_i (float): 土壤体积水含量,即某一特定深度处的土壤水含量。
            theta_s_i (float): 饱和土壤水含量,即土壤完全饱和时的水含量。
            theta_33_i (float): 土壤的田间持水量,即土壤中毛细孔隙所持有的水量,也是植物能轻易利用的水量。
            theta_1500_i (float): 永久凋萎点,即植物根系无法再吸收水分的土壤水含量。

            """

            Ai = (np.log(theta_33_i) - np.log(theta_1500_i)) / (np.log(1500) - np.log(33))

            # 计算土壤的饱和导水率 Ksi
            Ksi = 1930 * (theta_s_i - theta_33_i)**(3 - Ai)

            # 计算渗透系数 K
            K = Ksi * (theta_i / theta_s_i)**(3 + 2/Ai)

            return K

        # For the first and second layer:
        K1 = calc_permeability_k(theta[:,0], theta_s_layers, theta_33_layers, theta_1500_layers)
        K2 = calc_permeability_k(theta[:,1], theta_s_layers, theta_33_layers, theta_1500_layers)

        Sa1 = calc_SS(thetMAX_layers[0], theta[:,0], theta_1500_layers)
        Sb1 = calc_SS(thetMAX_layers[1], theta[:,1], theta_1500_layers)

         # Calculate the gradient of hydraulic conductivity
        delta_K = K1 - K2
        ln_Koa = np.log(K1)
        ln_Kob = np.log(K2)

        # if ln_Koa != ln_Kob:  # To avoid division by zero
        #     K_12 = delta_K / (ln_Koa - ln_Kob)
        # else:
        #     K_12 = 0
        if (ln_Koa != ln_Kob).any():  # To avoid division by zero when at least one element is different
            K_12 = delta_K / (ln_Koa - ln_Kob)
        else:
            K_12 = np.zeros_like(delta_K)  # Assuming delta_K is a numpy array, create an array of zeros with the same shape


        # Using Darcy's law to calculate Va_b for the first pair of layers
        Va_b1 = (K_12 * (Sb1 - Sa1) / ((100 / 1000) + ((theta_s_layers + theta_s_layers) / 2))) - 1
        self.Va_b1 = Va_b1 / 24
        
        # calculation for the second and third layer:
        K3 = calc_permeability_k(theta[:,2], theta_s_layers, theta_33_layers, theta_1500_layers)
        # Calculate the gradient of hydraulic conductivity
        delta_K = K2 - K3
        ln_Koa = np.log(K2)
        ln_Kob = np.log(K3)

        # if ln_Koa != ln_Kob:  # To avoid division by zero
        #     K_23 = delta_K / (ln_Koa - ln_Kob)
        # else:
        #     K_23 = 0
        if (ln_Koa != ln_Kob).any():  # To avoid division by zero when at least one element is different
            K_23 = delta_K / (ln_Koa - ln_Kob)
        else:
            K_23 = np.zeros_like(delta_K)  # Assuming delta_K is a numpy array, create an array of zeros with the same shape


        Sa2 = calc_SS(thetMAX_layers[1], theta[:,1], theta_1500_layers)
        Sb2 = calc_SS(thetMAX_layers[2], theta[:,2], theta_1500_layers)

        # Using Darcy's law to calculate Va_b for the second pair of layers
        Va_b2 = (K_23 * (Sb2 - Sa2) / ((150 / 1000) + ((theta_s_layers + theta_s_layers) / 2))) - 1
        self.Va_b2 = Va_b2 / 24


    # 计算土壤水库室中的水含量soil1_w、soil2_w、soil3_w;
    def calc_soilcontent(self):
        M_soil = np.array (self.meteo[['VWC_5cm_Avg', 'VWC_10cm_Avg', 'VWC_20cm_Avg']])
        M_soil[:, 0] *= 50 
        M_soil[:, 1] *= 100
        M_soil[:, 2] *= 150
        self.soil1w = M_soil[:, 0]   # 土壤中的水含量,kg/m2
        self.soil2w = M_soil[:, 1] 
        self.soil3w = M_soil[:, 2] 


    def calc_soilfx(self):
        self.calc_RSOIL ()
        self.calc_VDSO()
        self.calc_ESOIL()
        self.calc_soil_ab()
        self.calc_soilcontent()

class Calc_plant(Calc_air):
    '''
    calc_somefx ()  计算修正参数
    calc_st_res ()  计算气孔阻滞因子
    calc_VDPF ()    计算叶片沉降速率
    calc_ETRM()     计算蒸发速率
    calc_transOBT() 计算OBT生成速率
    calc_res()      调用本class中所有函数
    '''
    def __init__(self, pcse_results, start_date, end_date, plant_type, sta, HEG, x,drymatter, params):
        super ().__init__ (pcse_results, start_date, end_date, plant_type, sta, HEG, x, params)
        TAGP = pcse_results['TAGP']
        self.tagp = TAGP.values
        TWSO = pcse_results['TWSO']
        self.twso = TWSO.values
        PGASS = pcse_results['PGASS']    # 潜在同化速率, kg CH2O ha-1 day - 1
        self.Pco = PGASS.values        
        PMRES = pcse_results['PMRES']    # 潜在维持呼吸速率, kg CH2O ha-1 day - 1
        self.rd = PMRES.values

        self.calc_plantfx(pcse_results, plant_type,drymatter)

    # 计算 湿度加权函数,FE; 土层水含量加权函数, WATG;
    # 计算 温度加权函数, TTE; 辐射加权函数, RST;
    def calc_somefx(self, pcse_results, plant_type, VPD=0.2, WILTP=0.09):
        '''
        BODWA   初始土壤含水量,% 30%
        BODW    实际土壤含水量,mm
        WILTP   凋零点
        VPD     饱和蒸汽压差, 取0.2
        TMAXL   光合作用的温度上限, 取45℃
        TMINL   光合作用的温度下限, 取0 ℃
        TOPTL   光合作用的最适温度, 取25℃ optimal temperature for photosynthesis
        RCPAR   等于在气孔阻滞大于最小值两倍时测量的光合作用辐射通量密度
        PAR_net 净辐射平衡, W/m2
        '''
        # 计算湿度加权函数 DELTAE 和 FE
        DELTAE = (self.ROW - self.ROW * self.r_h) / 10.0   # DELTAE：蒸汽压差,hpa
        DELTAE = np.maximum (DELTAE, 0.001)  # 将DELTAE中小于0.001的值设为0.001
        self.FE = 1.0 - (VPD * DELTAE)
        self.FE = np.maximum (self.FE, 0.001)  # 将FE中小于0.001的值设为0.001

        # 计算温度加权函数 TTE
        TMINL = para_constant.Plant_dict[plant_type.upper ()]['TMINL']
        TOPTL = para_constant.Plant_dict[plant_type.upper ()]['TOPTL']
        TMAXL = para_constant.Plant_dict[plant_type.upper ()]['TMAXL']
        RCPAR = para_constant.Plant_dict[plant_type.upper ()]['RCPAR']
        TB = (TMAXL - TOPTL) / (TOPTL - TMINL)
        self.TTE = (self.TEMP - TMINL) / (TOPTL - TMINL) * \
                   ((TMAXL - self.TEMP) / (TMAXL - TOPTL)) ** TB
        self.TTE = np.maximum (self.TTE, 0.001)  # 将TTE中小于0.001的值设为0.001

    # 计算辐射加权函数 RST
        self.PAR = np.maximum (self.PAR, 0.001)  # 将PAR中小于0.001的值设为0.001
        self.RST = 1.0 + (RCPAR / self.PAR)

    # 土层水含量加权函数, WATG;
        BODWA = pcse_results['SM']
        self.BODWA = BODWA.values
        BODW1 = self.BODWA * 50   # BODWA = SM
        BODW2 = self.BODWA * 100
        BODW3 = self.BODWA * 150

        # 计算WAT1
        WAT1 = np.where (BODW1 > (WILTP * 50 + 2.5), 1,
                         np.where ((BODW1 > (WILTP * 50)) & (BODW1 < (WILTP * 50 + 2.5)),
                                   np.abs ((BODW1 - 2.5 - (WILTP * 50)) * 5.6) + 1, 15))

        # 计算WAT2
        WAT2 = np.where (BODW2 > (WILTP * 100 + 5), 1,
                         np.where ((BODW2 > (WILTP * 100)) & (BODW2 < (WILTP * 100 + 5)),
                                   np.abs ((BODW2 - 5.0 - (WILTP * 100)) * 2.8) + 1, 15))

        # 计算WAT3
        WAT3 = np.where (BODW3 > (WILTP * 150 + 7.5), 1,
                         np.where ((BODW3 > (WILTP * 150)) & (BODW3 < (WILTP * 150 + 7.5)),
                                   np.abs ((BODW3 - 7.5 - (WILTP * 150)) * 1.867) + 1, 15))

        # 根据条件计算self.WATG
        if self.ICOMP != 2:  # ICMP=2 表示牧草模型,1表示农业, 3 土豆, 4 小麦
            self.WATG = 0.4 * WAT1 + 0.3 * WAT2 + 0.3 * WAT3
        else:
            self.WATG = 0.4 * WAT1 + 0.6 * WAT2

    # 计算 植物-大气的总阻滞因子,RGES
    # 计算 植物传输阻力因子,RPHO, s/Cm
    def calc_st_res(self, ISTRE=1):
        # 计算 RCMINA 和 RC
        np.seterr (divide='ignore', invalid='ignore')  # 消除被除数为0的警告
        RCMINA = para_constant.RCMIN[self.ICOMP - 1] * np.ones_like(self.WATG)

        # 计算RC
        RC = 15. * RCMINA

        # 根据条件修改 RC1 的值
        condition = np.logical_or(self.PAR < 2, np.logical_and (self.PAR > 2, np.logical_or (self.rainfall > 0, self.r_h > 0.98))) | (ISTRE == 1)
        self.RC1 = np.where(condition, np.where(self.LAI != 0, RC / self.LAI, 1.E9), 1.E9)

        # 计算 RGES
        self.RGES = self.RAM + self.RB + self.RC1

        # 计算 RPHO
        self.RPHO = self.RST * self.WATG / (self.FE * self.TTE)
        self.RPHO = np.minimum (self.RPHO, 15)  # 保证 RPHO 不超过 15


    # 计算 达到平衡所需的时间常数 LAMPF, s-1
    # def calc_LAMPF(self):
    #     ALF = 1.1  # HTO 和水蒸气的特定浓度比
    #     A1 = (self.ROW / 1.E6) / (self.BWG * (self.RGES /100) * ALF)   
    #     self.LAMPF = np.maximum(A1, 6.2e-6)  # == 6.2e-5  *np.ones_like(resistance.VDSO


    # 计算 植物叶片的沉积速率VDPF,m s-1
    def calc_VDPF(self):
        # CONZ = 1.E-6     # 土豆和小麦中的最大氚浓度(与OBT的形成成比例)
        # ALF = 1.1        # HTO 和水蒸气的特定浓度比
        # self.ZEIT = 60   # 时间步长60s
        # A1 = (self.ROW * self.ZEIT * 1.E4) / (1.E6 * self.BWG * self.RGES * ALF)
        # A2 = np.exp (-np.minimum (A1, 50.))
        # self.VDPF = (((ALF / self.ROW) * CONZ * (1. - A2) * self.BWG) / self.ZEIT) / CONZ
        self.VDPF = 1 / (self.RGES)


    # 计算 植物蒸腾速率 ETRM 
    def calc_ETRM(self):
        """
        RELFEU (float): 相对湿度 (0-1)
        TEMP (float): 温度 (摄氏度)
        PAR (float): 光合有效辐射 (W/m^2)
        LEAFA (float): 叶面积指数 (m^2/m^2)
        """
        RATMOS = self.RAM + self.RB             # 大气阻滞因子之和,单位 m/s
        RSTOMA = self.RC1 * 100.                # 植物冠层气孔阻力因子,单位 m/s
        es = 0.6108 * np.exp(17.27 * self.TEMP / (self.TEMP + 237.3))  # 饱和蒸汽压,kpa,Tetens公式
        EFEUDE = (es - self.ea * self.r_h)         # 空气实际与饱和的蒸汽压之差,N m-2
        DELE = 4098 * es / (self.TEMP + 237.3)**2    # 饱和水汽压的温度依赖,单位 Pa
        
        GAM = 0.67                            # 大气干湿度压力系数
        RHO = 1.2923                          # 空气密度,单位 kg/m^3
        CP = 1.013                            # 比热容,单位 J/(kg K)
        LE = 2.45E+6                          # 水的蒸发潜热,单位 J/kg
        
        self.QSTR = self.PAR * (1- np.exp(-0.398 * self.LAI))
        ETRM = ((DELE * self.QSTR) + (RHO * CP * EFEUDE / RATMOS)) / (DELE + GAM * (1. + RSTOMA / RATMOS)) / LE
        self.ETRM = ETRM * 3600.              # 此处 植物蒸腾速率 ETRM kg m-2 h-1


    # 计算 植物实际光合作用率,HTO向OBT的转化 TROBT
    def calc_transOBT(self):
        '''
        BEWG. : 作物含水量,/mm2; 【不同作物含水量不同！！】
        RESP : 作物种类的总呼吸率, Rd
        PAKT0 : CO2同化吸收率,g CO2 /m2 /h
        PHMAX : 最大CO2同化吸收率,g CO2 /m2 /h
        PAKT : 作物整个冠层的同化吸收率, [即作物的质量转移速率,g CO2 /h]
        '''
        TEMPK = self.TEMP + 237.15
        C0 = [0.45, 1.0, 1.3, 1.7]             # 与作物种类有关的有机物质产生的值,mgCO2/M2/h
        PROD = [0.30, 0.48, 0.61, 0.7]         # 不同作物的平均干物质产量,h/g
        CO2DRY = 1.75                          # co2到干物质的转化因子,白天/夜晚权重因子(1.75/0.25)
        DISSOZ = 0.5                           # HTO分布耗散系数

        # drymatter_p = [160,170,600,500]      # 植物干物质, g m-2;

        # ALBEDO = 0.23      # 地表反射率
        # PHMAX = (0.158 * C0[self.ICOMP-1] * 1.0E9 * TEMPK * np.exp(-np.minimum(14200 / (TEMPK * 1.987), 50.))) / (
        #     1. + (np.exp(-np.minimum(47200 / (TEMPK * 1.987), 90.)) * np.exp(153.4 / 1.987)))

        # # PHMAX = (.044 * C0[self.ICOMP-1] * 1.0E9 * TEMPK * np.exp(-min (14200 / (TEMPK * 1.987), 50.))) / (
        # #         1. + (np.exp(-min(47200 / (TEMPK * 1.987), 90.)) * np.exp(153.4 / 1.987)))
        # PHMAX = PHMAX * 3.6
        # PAKT1 = (PHMAX + self.PAR * 0.036 * 0.6 * (1. - ALBEDO)) / (
        #         PHMAX + self.PAR * np.exp(-0.6 * self.LAI) * (1. - ALBEDO) * 0.036 * 0.6)
        
        # PAKT0 = PHMAX / 0.6 * np.log(PAKT1) / self.RPHO

        # RESP = 0.36 * PAKT0 + 1.67E-4 * self.tagp     # 根据公式的 
        # PAKT = PAKT0 - RESP 
        PAKT = (self.Pco -self.rd) * 1000 / 1e4 / 24 / 30 * 44
        PACT = PAKT * CO2DRY
        Tm = 2.0E-2 /24 
        self.TROBT =  PACT / PROD[self.ICOMP-1] * Tm * (1.75/0.25) * DISSOZ# / PROD[self.ICOMP-1]


    # 计算植物水库室中的水含量plant_w 和 有机物含量plant_o;kg/m2; 
    def calc_content(self,  drymatter):     
        self.plant_w = self.tagp / 1E4 * (1 - drymatter) / drymatter
        self.plant_wh = self.plant_w * 0.11
        self.plant_o = self.tagp / 1E4
        self.plant_oh = self.plant_o * 0.8 
        # 计算植物果实库室中的水含量friut_w 和 有机物含量friut_o;
        self.friut_w = self.twso / 1E4 * (1 - drymatter) / drymatter
        self.friut_wh = self.friut_w * 0.11   # 植物水中氢含量占水含量的11%；
        self.friut_o = self.twso /1E4 * 0.9
        self.friut_oh = self.friut_o * 0.8    # 植物有机质中氢含量占8%；

    def calc_plantfx(self,pcse_results,plant_type,drymatter):
        self.calc_somefx (pcse_results,plant_type)
        self.calc_st_res ()
        # self.calc_LAMPF ()
        self.calc_VDPF ()
        self.calc_ETRM()
        self.calc_transOBT()
        self.calc_content(drymatter)

from pcse.fileinput import CABOFileReader, YAMLCropDataProvider,YAMLAgroManagementReader, ExcelWeatherDataProvider
from pcse.util import WOFOST72SiteDataProvider
from pcse.base import ParameterProvider
from pcse.models import Wofost72_WLP_FD
from datetime import datetime, timedelta

def plantModel(plant_type, plantmodel, date, params):

    file_name = f"result_{uuid.uuid4()}.xlsx"
    for row in range(13, 1195):
        ws[f'B{row}'].value = float(params[2]) * 86.4
        ws[f'C{row}'].value = float(params[1])
        ws[f'D{row}'].value = float(params[1])
        ws[f'E{row}'].value = float(params[5])
        ws[f'F{row}'].value = float(params[0])
        ws[f'G{row}'].value = float(params[3])
    
    wb.save(os.path.join(data_dir, 'meteo', file_name))

    # 设置气象参数
    weatherfile = os.path.join(data_dir, 'meteo', file_name)
    wdp = ExcelWeatherDataProvider(weatherfile)
    
    if plant_type.upper() == 'GRASS':
        # 设置植物生理参数
        cropfile = os.path.join(data_dir, 'crop', 'Grass.crop')
        cropd = CABOFileReader(cropfile)
        # 设置作物生长收获相关参数
        agromanagement_file = os.path.join (data_dir, 'agro', "Grass.agro")
        agromanagement = YAMLAgroManagementReader(agromanagement_file)
    else:
        cropd = YAMLCropDataProvider()
        # plantmodel_EXP = [crop, available_varieties, soilfile_name, agrofile_name]
        cropd.set_active_crop(plantmodel[0], plantmodel[1])
        agromanagement_file = os.path.join(data_dir, 'agro', plantmodel[3])
        agromanagement = YAMLAgroManagementReader(agromanagement_file)

# 初始化参数，运行wofost
    parameters = ParameterProvider(cropdata=cropd, soildata=soild, sitedata=sited)
    wofsim = Wofost72_WLP_FD(parameters, wdp, agromanagement)

    start_date = datetime.strptime(date[0], '%Y-%m-%d').date()
    end_date = datetime.strptime(date[1], '%Y-%m-%d').date()
    wofsim.run_till(end_date)
    # wofsim.run_till_terminate()
    # C:\Users\auror\miniconda3\envs\cttm\Lib\site-packages\pcse\conf\Wofost72_WLP_FD.conf
    df_results = pd.DataFrame(wofsim.get_output())
    
    # 输出模拟日期内所需的参数值
    first_date = df_results.iloc[0]['day']
    delta = first_date - start_date
    day_delta = delta.days
    df_results = df_results.set_index("day")
    df_results = df_results.iloc[abs(day_delta):]
    
    # 使用线性插值将每日的数据转变为每小时数据
    df_results.index = pd.to_datetime(df_results.index)
    df_resampled = df_results.resample('H').interpolate(method='linear')

    # df_results.to_csv(f'./data/pcse_output/{plant_type}_pcse_results_daily.cvs')
    # df_resampled.to_csv(f'./data/pcse_output/{plant_type}_pcse_results_hourly.cvs')
    os.remove(weatherfile)
    return df_resampled

def model_function(params):
    plantmodel = ['potato', 'Potato_702', 'ec2.soil', 'potato.agro']
    plant_type = 'ROOT_VEG'

    start_date = '2021-06-20'
    end_date = '2021-09-20'
    date = [start_date, end_date]
    stability = 1
    HEG = 20
    x = 1000

    char='HTO'
    drymatter = 0.21
    pcse_results = plantModel(plant_type, plantmodel, date, params)

    airfx =  Calc_air(pcse_results, start_date, end_date, plant_type, stability, HEG, x, params)
    soilfx = Calc_soil(pcse_results, start_date, end_date, plant_type, stability, HEG, x, params)
    plantfx = Calc_plant(pcse_results, start_date, end_date, plant_type, stability, HEG, x,drymatter, params)

    k_array = transfer_rates_baomi(char, plant_type, airfx, soilfx, plantfx)
    As = solve_con('HTO', 9.78728318e+08, k_array, LAMBDA_T_H)

    return As[2160]

#%%
# Problem definition
problem = {
    'num_vars': 9, # Number of parameters
    'names': ['WS10m_avg', 'Ta_Avg', 'DR_Avg', 'Rain_Tot', 'RH_Avg', 'Pvapor_Avg', 'VWC_5cm_Avg', 'VWC_10cm_Avg', 'VWC_20cm_Avg'],
    'bounds': [[0.1, 10], [-5, 38], [20, 300], [0, 5], [50, 95], [0.1, 3], [0.1, 0.4], [0.1, 0.4], [0.1, 0.4]]
}
# Generate samples
param_values = sample(problem, N=1000, num_levels=4, optimal_trajectories=None)
print(param_values.shape)

data_dir = os.path.join(os.getcwd(), "./data")
# 设置土壤相关参数
soilfile = os.path.join(data_dir, 'soil', 'ec2.soil')

soild = CABOFileReader(soilfile)

# 设置站点参数
sited = WOFOST72SiteDataProvider(WAV=10)

pcsefile = os.path.join(data_dir, 'meteo', 'meteo_usefor_pcse.xlsx')
wb = load_workbook(pcsefile)

ws = wb.active
def evaluate_model(params):
    return model_function(params)

# %%
num_cores = -1
Y = Parallel(n_jobs=num_cores)(delayed(evaluate_model)(params) for params in tqdm(param_values))


# %%
